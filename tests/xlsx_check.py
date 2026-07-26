# -*- coding: utf-8 -*-
"""Независимый валидатор .xlsx: python стандартной библиотекой.

Аргумент: путь к файлу. Проверяет целостность zip (testzip + CRC при чтении)
и корректность каждого XML-члена (xml.etree выбрасывает при любой ошибке
разметки). Если установлен openpyxl — дополнительно читает лист и возвращает
число непустых строк. stdout: JSON {ok, members, openpyxl?, error?}.
"""
import sys, json, zipfile
import xml.etree.ElementTree as ET


def main(path):
    res = {"ok": False, "members": 0}
    try:
        with zipfile.ZipFile(path) as z:
            bad = z.testzip()
            if bad is not None:
                res["error"] = "битый член: " + bad
                print(json.dumps(res)); return
            names = z.namelist()
            for req in ("[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"):
                if req not in names:
                    res["error"] = "нет обязательного члена: " + req
                    print(json.dumps(res)); return
            for n in names:
                data = z.read(n)  # чтение сверяет CRC
                if n.endswith(".xml") or n.endswith(".rels"):
                    ET.fromstring(data)
            res["members"] = len(names)
        res["ok"] = True
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = sum(1 for row in ws.iter_rows(values_only=True)
                       if any(v is not None for v in row))
            res["openpyxl"] = rows
        except ImportError:
            pass
    except Exception as e:  # любой сбой — невалидный файл
        res["error"] = "%s: %s" % (type(e).__name__, e)
    print(json.dumps(res))


if __name__ == "__main__":
    main(sys.argv[1])
